"""Example for generating pattern coverage maps."""
import os
import re
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.colors import LinearSegmentedColormap, TwoSlopeNorm
from openpyxl import load_workbook, Workbook

import antenna_analysis
import antenna_analysis.data_load_store


def make_new_color_settings(CENTER, margin: float = 0):
    """Make a new color setting to be passed to the contour plot."""
    if margin == 0:
        vmin = None
        vmax = None
        lin_range = (0.75, 0.75)
    else:
        vmin = CENTER - margin
        vmax = CENTER + margin
        lin_range = (0.5, 0.9)
    colors_red = np.flip(plt.cm.Reds(np.linspace(*lin_range, 256)), axis=0)
    colors_green = plt.cm.Greens(np.linspace(*lin_range, 256))
    all_colors = np.vstack((colors_red, colors_green))
    cmap = LinearSegmentedColormap.from_list('mycmap', all_colors)
    divnorm = TwoSlopeNorm(vmin=vmin, vcenter=CENTER, vmax=vmax)
    return cmap, divnorm


def store_ccdf_data(figure):
    """Export the CCDF data to XLS."""
    if not os.path.exists('output'):
        os.makedirs('output')
    filename = 'ccdf_data'
    filename = os.path.join('output', filename+'.xlsx')
    if os.path.isfile(filename):
        workbook = load_workbook(filename)
    else:
        workbook = Workbook()

    for line in range(len(figure.axes[0].get_lines())):
        full_label = f'Trace_#_{line}'
        freq_label = f'Trace_#_{line}'
        if figure.axes[0].get_legend() is not None:
            full_label = figure.axes[0].get_legend().get_texts()[line].get_text()
            freq_label = re.search('--> (.*) @Port', full_label).group(1)
        else:
            freq_label = f'Trace_#_{line}'
        worksheet = workbook.create_sheet(title=f'{line}_{freq_label}')
        x_data = figure.axes[0].get_lines()[line].get_xdata()
        y_data = figure.axes[0].get_lines()[line].get_ydata()

        # step = 0.05
        # values = np.flip(np.arange(0, 1+step, step))

        step = 1
        values = np.append(np.insert(np.arange(np.ceil(x_data.min()),
            np.floor(x_data.max())+step, step), 0, x_data.min()), x_data.max())

        new_x_data = np.array([])
        new_y_data = np.array([])
        for ind, val in enumerate(values):
            array = np.asarray(x_data)
            data_idx = (np.abs(array - val)).argmin()
            new_x_data = np.append(new_x_data, x_data[data_idx])
            new_y_data = np.append(new_y_data, y_data[data_idx])

        # Write the full label as header
        worksheet.cell(1, 1, full_label)
        worksheet.cell(2, 1, 'X Data')
        worksheet.cell(2, 2, 'Y Data')
        for ind, x in enumerate(new_x_data):
            worksheet.cell(ind+3, 1, x)
            worksheet.cell(ind+3, 2, new_y_data[ind])

    workbook.save(filename)


# Initialize an object from the class
DATA = antenna_analysis.Patterns()

# For this example load a half lambda dipole at 1.5 GHz and
# a patch on the same frequency.
DATA.generate_analytical_pattern()  # Dipole
DATA.generate_analytical_pattern({'ant_type': 'circular patch'})  # Circ. patch

# Print the list of loaded Patterns
PATT_LIST = DATA.list_patterns()
print('Patterns List:')
print("\n".join(PATT_LIST))
print('\n')

# in this example we work with Gabs but it can also be GL or GR or any other
# coverage required
FIELD_TO_ANALYZE = 'Gabs'
PATTERNS_TO_MANIPULATE = [0, 1]

PLOT_FIELD, PLOT_PHASE, LEGENDS = \
    DATA.prepare_for_plotting(PATTERNS_TO_MANIPULATE,
                              FIELD_TO_ANALYZE)
# Creates the plotting object
PLOTS = antenna_analysis.PatPlot(PLOT_FIELD, LEGENDS)

PLOTS.set_plot_params({'dynamic_range': ['auto', 40, 'log'],
                       'plot_range': [[0, 180], [0, 360]],
                       'plot_type': 'linear',
                       'output_path_pad': 'example_contour'
                       })

# create new colormap and normalization around the required antenna gain value.
cmap, divnorm = make_new_color_settings(CENTER=0, margin=3)

# Customizations of the plots - see _parse_customizations for details
customizations = {
    "title": "Example title",
    "xlabel": "Example XLABEL",
    "ylabel": "Example YLABEL",
    "show legend": False,
    "set Colormap": cmap,
    'set Normalization': divnorm
    }

PLOTS.make_countour_plot(0, customizations=customizations)
PLOTS.make_countour_plot(1, customizations=customizations)

PLOTS.set_plot_params({'output_path_pad': 'statistics',
                       'dynamic_range': ['MaxFull3D', 40, 'log'],
                       'plot_range': [[0, 180], [0, 360]],
                       })
ccdf_fig = PLOTS.make_statistics_plot(stat_plot_type='CCDF',
                                      customizations=customizations)

store_ccdf_data(ccdf_fig)

# Complete the script
print('\n')
print('Done')
