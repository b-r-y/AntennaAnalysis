'''
Created on Nov 9, 2018

@author: bry
'''
import sys
import os
import ast
import json
import copy

import csv

from collections import OrderedDict
from openpyxl import load_workbook
from openpyxl import Workbook
import numpy as np
import includes


def get_p_struct():
    '''
    Static method used to initialize a new empty ``patterns`` list member. This ordered
    dictionaty describes how patterns are stored internally and what information is kept at
    import.

    Arguments:
        source (str): a string typically containing the folder the imported filed were stored
        file (str): a string with the file name of the pattern(s) imported
        frequency (float): the pattern frequency. This is extracted from the file name in case
            of CST files
        port (str): antenna port identification - currently unused
        th_step (int): Theta angle stepping for the pattern resolution. This is always integer
            and constant for the entire pattern.
        ph_step (int): Phi angle stepping for the pattern resolution. This is always integer
            and constant for the entire pattern.
        ff (array): ff stand for far frield. The `ff` stores the electric field in E_theta and
            E_Phi components. The dimensions of `ff` are 2 x n_theta x n_phi where n_theta/n_phi
            are the total number of samples in Theta and Phi, spherical coordinate system as
            defined in [IEEESTDTest]_. E_theta is stores as index 0 and E_phi as index 1.
    '''
    p_struct = OrderedDict()
    p_struct['source'] = [""]
    p_struct['file'] = [""]
    p_struct['frequency'] = []
    p_struct['port'] = []
    p_struct['th_step'] = []
    p_struct['ph_step'] = []
    p_struct['rot_offset'] = [0,0,0]
    # Default Far Field stored as Real/Imaginary Electric Fields in Theta/Phi\
    # Linear polarization
    p_struct['ff'] = []
    return p_struct

def check_load_data_input(input_type, input_file_or_folder):
    '''
    Check the loading data inputs and returns an error if input type and required file/folder types
    do not match.
    '''
    if input_type == 'json' or input_type == 'AAU_Legacy' or input_type == 'AAU_Satimo' or \
        input_type == 'CST_File' or input_type == 'AMS32':
        if not os.path.isfile(input_file_or_folder):
            sys.exit("{} file loading type requires file path input!".format(input_type))
    elif input_type == 'CST_Folder' or input_type == 'CST_Par_Sweep':
        if not os.path.isdir(input_file_or_folder):
            sys.exit("{} loading type requires folder path input!".format(input_type))
    else:
        sys.exit("Unknown import type!")

def load_from_json(patterns_obj, filename):
    '''
    This method imports multiple patterns from a single JSON file. Must be generated with this
    class for compatibility.

    Arguments:
        file (str): full path to the file.

    Returns:
        ff (array): appends multiple ``patterns`` to the list.
    '''

    file_path = open(filename, 'rt')
    json_data = json.load(file_path, object_pairs_hook=OrderedDict)
    file_path.close()

    patterns_obj.__dict__.update(json_data)

    for i, subarray in enumerate(patterns_obj.patterns):
        sub_list = []
        for subsubarray in subarray['ff']:
            sub_list.append([ast.literal_eval(row) for row in subsubarray[1]])

        if sub_list.__len__() == 0:
            far_field = []
        else:
            e_th = np.array(sub_list[0])
            e_ph = np.array(sub_list[1])
            far_field = np.stack((e_th,e_ph), axis=0)

        patterns_obj.patterns[i]['ff'] = far_field
    return patterns_obj


def load_from_ams32(patterns_obj, filename):
    data = []
    with open(filename, newline='') as csvfile:
        spamreader = csv.reader(csvfile, delimiter=',')
        for row in spamreader:
            data.append(row)

    start_idx = []
    hor_idx = []
    ver_idx = []
    end_idx = []
    span = []
    for idx, row in enumerate(data):
        if row:
            if row[0] == 'Polarization':
                print('Found data start at line {}'.format(idx))
                start_idx = idx
            if row[0] == 'Horizontal':
                print('Found horizontal start at line {}'.format(idx))
                hor_idx = idx
            if row[0] == 'Vertical':
                print('Found vertical start at line {}'.format(idx))
                ver_idx = idx
            if row[0] == 'Point Values':
                print('Found data end at line {}'.format(idx))
                end_idx = idx
    # Find the sizes and angular steps
    for idx, row in enumerate(data[hor_idx+1:]):
        if not row:
            span = idx
            print('span = {}'.format(span))
            break

    no_freqs = int((ver_idx - hor_idx - 2)/(2*span + 2))

    for f in range(0, no_freqs, 1):
        # Process the clumps of data one by one
        new_pat = __ams32_single_table(data[ver_idx + 1 + (f*(2*span+2)):\
                                            ver_idx+3+2*span + (f*(2*span+2))]\
                                            ,\
                                            data[hor_idx+1 + (f*(2*span+2)):\
                                            hor_idx+3+2*span + (f*(2*span+2))]\
                                            , span)


        new_pat['source'] = os.path.dirname(filename)
        new_pat['file'] = os.path.basename(filename)
        patterns_obj.patterns.append(new_pat)

def __ams32_single_table(vdata, hdata, span):
    v_power = []
    h_power = []
    v_phase = []
    h_phase = []
    th_step = int(round(float(vdata[3][2]))) - int(round(float(vdata[2][2])))
    NoThetaSamples = int((180+th_step)/th_step)
    ph_step = int(round(float(vdata[0][4]))) - int(round(float(vdata[0][3])))
    NoPhiSamples = int((360)/ph_step)

    e_th = np.zeros((NoThetaSamples, NoPhiSamples), dtype=complex)
    e_ph = np.zeros((NoThetaSamples, NoPhiSamples), dtype=complex)
    for i in range(2, span):
        h_power.append([float(val) for val in hdata[i][3:]])
        v_power.append([float(val) for val in vdata[i][3:]])
    for i in range(3+span, (2*span)+1):
        h_phase.append([float(val) for val in hdata[i][3:]])
        v_phase.append([float(val) for val in vdata[i][3:]])
    v_power = np.array(v_power)
    v_phase = np.array(v_phase)
    h_power = np.array(h_power)
    h_phase = np.array(h_phase)

    e_th[:-1, :] = (10**(v_power[:, :-1]/20)) * \
                    np.cos(v_phase[:, :-1]*np.pi/180) + \
                    (1j * (10**(v_power[:, :-1]/20)) * \
                    np.sin(v_phase[:, :-1]*np.pi/180))
    e_ph[:-1, :] = (10**(h_power[:, :-1]/20)) * \
            np.cos(h_phase[:, :-1]*np.pi/180) + \
            (1j * (10**(h_power[:, :-1]/20)) *
             np.sin(h_phase[:, :-1]*np.pi/180))
    far_field = np.stack((e_th, e_ph))

    new_pat = get_p_struct()
    new_pat['frequency'] = float(vdata[0][1])*1E6
    new_pat['port'] = 1
    new_pat['th_step'] = th_step
    new_pat['ph_step'] = ph_step
    new_pat['ff'] = far_field

    return new_pat


def load_from_aau_legacy(patterns_obj, filename):
    '''
    Warning:
        This method has hardcoded calibration data for the X-Band project!

    Todo:
        Make generic calibration data loading and header treatment.

    This method is used to load data from the old AAU anechoic chamber captured with the
    Keysight VNA using triggers.

    Arguments:
        filename (str): full path to the file with the CSV data

    Returns:
        far_field (array): appends multiple ``new_pat`` to the list.
    '''
    # Primary input constants
    file_header_lines = 5
    freq_points = 41
    meas_header_lines = 2
    meas_footer_lines = 3
    port = 1
    th_step = 5
    ph_step = 5
    th_range = [-90, 90]
    ph_range = [0, 180]
    ref_ant = np.array([11,11,11,11,11,11,11,11,11,11,11,11.5,11.5,11.5,11.5,11.5,11.5,11.5,\
                        11.5,11.5,11.5,11.6,11.67,11.74,11.81,11.88,11.95,12.02,12.09,12.16,\
                        12.23,12.3,12.37,12.44,12.51,12.58,12.65,12.72,12.79,12.86,12.93])

    # Calculated values
    th_samples = len(range(th_range[0], th_range[1], th_step))+1
    ph_samples = len(range(ph_range[0], ph_range[1], ph_step))+1

    v_mag =[]
    v_ph = []
    h_mag =[]
    h_ph = []
    skip = file_header_lines + meas_header_lines
    for i in range(0,th_samples*ph_samples):
        if not i==0:
            skip = skip + (freq_points + meas_footer_lines + meas_header_lines)
        freq, t_v_mag, t_v_ph, cal_v_mag, cal_v_ph, t_h_mag, t_h_ph, cal_h_mag, cal_h_ph = \
            np.genfromtxt(filename, skip_header=skip, delimiter=',', \
                          usecols=(0,1,2,3,4,5,6,7,8),max_rows=freq_points, unpack=True)
        v_mag.append(t_v_mag - cal_v_mag + ref_ant)
        v_ph.append(t_v_ph - cal_v_ph)
        h_mag.append(t_h_mag - cal_h_mag + ref_ant)
        h_ph.append(t_h_ph - cal_h_ph)

    v_mag = np.array(v_mag)
    v_ph = np.array(v_ph)
    h_mag = np.array(h_mag)
    h_ph = np.array(h_ph)
    e_th_all = (10**(h_mag/20)) * np.cos(h_ph*np.pi/180) + \
                                                (1j * (10**(h_mag/20))*np.sin(h_ph*np.pi/180))
    e_ph_all = (10**(v_mag/20)) * np.cos(v_ph*np.pi/180) + \
                                                (1j * (10**(v_mag/20))*np.sin(v_ph*np.pi/180))

    for i in range(0, freq.__len__()):
        new_pat = get_p_struct()
        new_pat['source'] = os.path.dirname(filename)
        new_pat['file'] = os.path.basename(filename)
        new_pat['frequency'] = freq[i]
        new_pat['port'] = port
        new_pat['th_step'] = th_step
        new_pat['ph_step'] = ph_step

        e_th = np.zeros((th_samples, 2*(ph_samples-1)), dtype=complex)
        e_ph = np.zeros((th_samples, 2*(ph_samples-1)), dtype=complex)

        e_th_t = np.reshape(e_th_all[:, i], (th_samples, ph_samples), 'F')
        e_ph_t = np.reshape(e_ph_all[:, i], (th_samples, ph_samples), 'F')

        e_th[0:(th_samples-1)//2 + 1, 0:th_samples-1] = e_th_t[(th_samples-1)//2:, :-1]
        e_ph[0:(th_samples-1)//2 + 1, 0:th_samples-1] = e_ph_t[(th_samples-1)//2:, :-1]

        e_th[0:(th_samples-1)//2 + 1, th_samples-1:] = \
            np.flipud(e_th_t[0:(th_samples-1)//2 + 1, :-1])
        e_ph[0:(th_samples-1)//2 + 1, th_samples-1:] = \
            np.flipud(e_ph_t[0:(th_samples-1)//2 + 1, :-1])

        far_field = np.stack((e_th, e_ph))
        new_pat['ff'] = far_field

        patterns_obj.patterns.append(new_pat)

    return patterns_obj

def load_from_aau_satimo(patterns_obj, filename):
    '''
    This method imports Satimo measured files. Satimo files can contain multiple patters at
    multiple frequencies. The import uses the electric fields exported by the Satimo SW.

    Todo:
        Make generic import using any of the possible Satimo exports and not only the E fields

    Warning:
        Satimo exports seem to flip the labeling of the E_Theta and E_Phi!

    Arguments:
        filename (str): full path to the file with the TXT data

    Returns:
        ff (array): appends multiple ``new_pat`` to the list.
    '''

    port = 1

    freq, phi, theta, eph_r, eph_i, eth_r, eth_i = \
        np.genfromtxt(filename, skip_header=2, delimiter='\t', \
                      usecols=(0,1,2,3,4,5,6), unpack=True)

    # ATTENTION!!!! There is something wrong with the SATIMO export!!!!!
    e_ph_all = eth_r + 1j*eth_i
    e_th_all = eph_r + 1j*eph_i

    freq = np.unique(freq)
    t_th = np.unique(theta)*180/np.pi
    t_ph = np.unique(phi)*180/np.pi
    th_range  = np.array([int(round(t_th[0])), int(round(t_th[-1]))])
    ph_range  = np.array([int(round(t_ph[0])), int(round(t_ph[-1]))])

    th_step = int(round(t_th[-1])) - int(round(t_th[-2]))
    ph_step = int(round(t_ph[-1])) - int(round(t_ph[-2]))

    th_samples = len(range(th_range[0], th_range[1], th_step))+1
    ph_samples = len(range(ph_range[0], ph_range[1], ph_step))+1

    new_th_samples = len(range(0,180, th_step))+1
    new_ph_samples = len(range(0,360, ph_step))

    e_th_all = np.reshape(e_th_all, (th_samples, ph_samples, freq.__len__()), 'F')
    e_ph_all = np.reshape(e_ph_all, (th_samples, ph_samples, freq.__len__()), 'F')

    for i in range(0, freq.__len__()):
        new_pat = get_p_struct()
        new_pat['source'] = os.path.dirname(filename)
        new_pat['file'] = os.path.basename(filename)
        new_pat['frequency'] = freq[i]
        new_pat['port'] = port
        new_pat['th_step'] = th_step
        new_pat['ph_step'] = ph_step

        e_th = np.zeros((new_th_samples, new_ph_samples), dtype=complex)
        e_ph = np.zeros((new_th_samples, new_ph_samples), dtype=complex)

        e_th[:,0:int(new_ph_samples//2)] = e_th_all[int((th_samples-1)//2):,:,i]
        e_ph[:,0:int(new_ph_samples//2)] = e_ph_all[int((th_samples-1)//2):,:,i]

        e_th[:,int(new_ph_samples//2):] = np.flipud(e_th_all[0:int((th_samples-1)//2)+1,:,i])
        e_ph[:,int(new_ph_samples//2):] = np.flipud(e_ph_all[0:int((th_samples-1)//2)+1,:,i])

        new_pat['ff'] = np.stack((e_th, e_ph))

        patterns_obj.patterns.append(new_pat)

    return patterns_obj

def load_from_cst_file(patterns_obj, filename):
    '''
    This method imports individual CST simulated files. Each file contains a single radiation
    pattern.

    Since CST exported patters do NOT include the frequency or the port in the file certain
    default CST settings are needed so that the default export from the post processing template
    generates a file name that can be parsed and these parameters can be extracted. See the CST
    settings section of this guide for details.

    Todo:
        Make generic import using any of the possible Satimo exports and not only the E fields

    Arguments:
        filename (str): full path to the file with the TXT data

    Returns:
        ff (array): appends single ``new_pat`` to the list.
    '''
    new_pat = get_p_struct()
    new_pat['source'] = os.path.dirname(filename)
    new_pat['file'] = os.path.basename(filename)

    frequency = float(filename[filename.find("(f=")+3 : filename.find(")")])
    if not frequency:
        sys.exit("Something went wrong with the file name parsing of the frequency... \
            check the format. Trying to parse: {}".format(filename))
    if frequency < 100: # assumes frequency in GHz. Otherwise assumed in MHz
        frequency = frequency*1E9 # assumes frequency in GHz.
    else:
        frequency = frequency*1E6 # Otherwise assumed in MHz
    new_pat['frequency'] = frequency

    port = (filename[filename.find("[")+1 : filename.find("]")])
    if not port:
        sys.exit("Something went wrong with the file name parsing of the port... \
            check the format. Trying to parse: {}".format(filename))
    new_pat['port'] = port

    cst_file = open(filename,"r")
    t_header = cst_file.readline()
    cst_file.close()

    # Manipulate the header to extract necessary information and re-format appropriately
    t_header = t_header.replace(" ","") # remove all white spaces
    t_header = t_header.split(']') #split along the closing dimension bracket
    t_header = t_header[:-1] # remove the last \n symbol

    header = []
    for i in range(t_header.__len__()): # add back the closing bracket removed in the split
        header.append(t_header[i]+"]")

    ## The field_option is currently not used
    t_field_option = header[2][header[2].find('(')+1 : header[2].find(')') ]
    if t_field_option == "Grlz":
        field_option = "Gr"
    elif t_field_option == "Gain":
        field_option = "G"
    else:
        sys.exit('Only importing Realized Gain {} supported for now - \
            change the export settings in CST'.format(field_option))

    t_field_units = header[3][header[3].find('[')+1 : header[3].find(']') ]

    if t_field_units in ('dB', 'dBi'):
        field_units = "dB"
    elif t_field_units == '':
        field_units = "MA"
    else:
        sys.exit('Only importing in dB and MA format supported for now - \
            change the export settings in CST')

    t_polarization = header[3][header[3].find('(')+1 : header[3].find(')') ]
    if t_polarization == "Left":
        polarization = "circular"
    elif t_polarization == "Theta":
        polarization = "linear"
    else:
        sys.exit('Only importing in circular and linear polarization supported for now - \
            change the export settings in CST')

    theta_cst, phi_cst, pol1_1, pol1_2, pol2_1, pol2_2 = \
        np.genfromtxt(filename, skip_header=2,usecols=(0,1,3,4,5,6),unpack=True)

    th_step = int(theta_cst[1] - theta_cst[0])
    new_pat['th_step'] = th_step# determine the Theta step
    ph_step = int(phi_cst[int(180//th_step + 1)] - phi_cst[0])
    new_pat['ph_step'] = ph_step

    pol1_1 = np.reshape(pol1_1, [180//th_step +1,360//ph_step], 'F')
    pol1_2 = np.reshape(pol1_2, [180//th_step +1,360//ph_step], 'F')
    pol2_1 = np.reshape(pol2_1, [180//th_step +1,360//ph_step], 'F')
    pol2_2 = np.reshape(pol2_2, [180//th_step +1,360//ph_step], 'F')

    #goes here if the imported units are in dB so it can convert to linear E field magnitude
    if field_units =='dB':
        pol1_1 = 10**(pol1_1/20)
        pol2_1 = 10**(pol2_1/20)
    elif field_units =='MA':
        pol1_1 = np.sqrt(pol1_1)
        pol2_1 = np.sqrt(pol2_1)
    else:
        sys.exit('Unknown field units!!!')

    # Convert the phases to radians
    pol1_2 = pol1_2*np.pi/180
    pol2_2 = pol2_2*np.pi/180

    # Calculate the complex E fields
    # This is Theta in the case of linear and Left in the case of circular
    pol1_e = pol1_1*np.cos(pol1_2) + (1j * pol1_1*np.sin(pol1_2))
    # This is Phi in the case of linear and Right in the case of circular
    pol2_e = pol2_1*np.cos(pol2_2) + (1j * pol2_1*np.sin(pol2_2))

    if polarization == 'circular': # Convert to linear in case of circular import
        e_th = (1/np.sqrt(2))*(pol1_e + pol2_e) # This becomes Theta
        e_ph = ((1j)/np.sqrt(2))*(pol1_e - pol2_e) # This becomes Phi
    else:
        e_th = pol1_e
        e_ph = pol2_e

    new_pat['ff'] = np.stack((e_th, e_ph))

    patterns_obj.patterns.append(new_pat)

    return patterns_obj

def load_from_cst_folder(patterns_obj, folder):
    '''
    This method imports multiple CST simulated files_found located in the Export folder.
    Each file contains a single radiation pattern.

    Same guidelines for CST settings apply.

    Arguments:
        folder (str): full path to the folder with the exported patterns

    Returns:
        ff (array): appends multiple ``patterns`` to the list.
    '''
    files_found = [file for file in os.listdir(folder) if \
                   os.path.isfile(os.path.join(folder, file))]

    for file in files_found:
        patterns_obj = load_from_cst_file(patterns_obj, os.path.join(folder,file ))

    return patterns_obj

def load_from_cst_par_sweep(patterns_obj, folder):
    '''
    This method imports multiple CST simulated files located in the Cashed folder. It traverses
    the default CST parameter sweep structure and loads all files found. Each file contains a
    single radiation pattern.

    Same guidelines for CST settings apply.

    Arguments:
        folder (str): full path to the folder with the top level parameter sweep folder.

    Returns:
        ff (array): appends multiple ``patterns`` to the list.
    '''

    cst_project_name = folder[folder.rfind('\\')+1:]
    folder = os.path.join(folder, 'Result', 'Cache')
    rundirs = [f for f in os.listdir(folder) if os.path.isdir(os.path.join(folder, f))]

    for r_dir in rundirs:
        current_run_folder = os.path.join(folder, r_dir)
        only_dirs_in_run_folder = \
            [f for f in os.listdir(current_run_folder) \
             if os.path.isdir(os.path.join(current_run_folder, f))]
        # always picks the last pass if multiple exist
        current_project_pass_folder = only_dirs_in_run_folder[-1]
        if only_dirs_in_run_folder[-1] != cst_project_name:
            current_project_pass_folder = os.path.join(current_project_pass_folder, \
                                                       cst_project_name)
        working_folder = os.path.join(current_run_folder, current_project_pass_folder, \
                                      'Export', 'Farfield')

        print("Loading data from {}".format(working_folder))
        patterns_obj = load_from_cst_folder(patterns_obj, working_folder)

    return patterns_obj

def save_to_json(patterns_obj, filename):
    '''
    Save a patterns structure to a JSON file.

    Arguments:
        filename(str): full path to the filename to be used.
    '''
    # We don't want to change the original verison of the class, so make a deep copy
    serialized_self = copy.deepcopy(patterns_obj)

    # To make the json file more readable on these huge arrays, we make the subarrays
    # into strings, so the python json module automatically inserts newlines.

    # Dump serialized version to a file.
    file_path = open(filename, 'wt')

    for i, subarray in enumerate(patterns_obj.patterns):
        sub_list = []
        to_store = []
        for i_i, subsubarray in enumerate(subarray['ff']):
            if i_i == 0:
                sub_list_header = "E Theta"
            else:
                sub_list_header = "E Phi"
            sub_list = [sub_list_header]
            sub_list.append([str(row.tolist()) for row in subsubarray])
            to_store.append(sub_list)
            serialized_self.patterns[i]['ff'] = to_store

    order = ['comment_static_vars', 'version', 'comment_auto_populated', 'patterns']

    ord_dict = OrderedDict()
    for key in order:
        ord_dict[key] = serialized_self.__dict__[key]

    json.dump(ord_dict, file_path, indent=2)
    file_path.close()

def get_analytical_parameters(input_parameters = None):
    '''
    Defines defaults for analytical pattern generation and overwrites those defaults if values are
    submited.

    Args:
        input_parameters (dict): a python dictionary containing the various parameters. The \
            supported keys and values are defined below:

            - ant_type (:obj:`str`): Defines the analytical antenna type to be generated. Those can\
                                    be:

                - *'dipole'* - finite length analytical dipole - default

                - 'point source' - ideal isotropic radiator

                - 'delta function' - pencil beam is a single direction.

                - 'helix' - a helical antenna according to [BALANIS]_

            - resolution (:obj:`int`, :obj:`int`): defines the [theta,phi] resolution in \
                    degrees - [1,1] as default
            - frequency (:obj:`int`): used mostly for the dipole type and the wavelength scaling \
                    in Hz. Default is 1E9 for 1GHz dipole.
            - length (:obj:`float`): finite dipole length as a fraction of wavelength - \
                    dimenssionless. Default is 0.5 for a half-lambda dipole.
            - efficiency (:obj:`int`): once generated the analytical dipole is scaled to certain \
                        efficiency defined in percent. Default is 100.
            - polarization (:obj:`str`): a string defining the polarization for helix. Options \
                        are:

                - *'RHCP'* - right hand polarization - default

                - 'LHCP' - left hand polarization

            - circ2lambda_ratio (:obj:`float`): The ration between the helix circumference and the \
                wavelength. For optimal helix use the default 1.
            - step2lmabda_ratio (:obj:`float`): The ration between the helix turn step and the \
                wavelength. For optimal helix use the default 0.23.
            - no_turn (:obj:`int`): number of turns for the helix - default is 5
            - helix_type (:obj:`str`): choose between *'Hansen-Woodyard'* or 'ordinary' \
                helix types. Default is *'Hansen-Woodyard'*.
    Returns:
        dict: A dictinary of updated parameters for analytical antenna generation
    '''
    # State the defaults
    parameters = {  'ant_type':'dipole',
                    'resolution':[1, 1],
                    'frequency':1E9,
                    'length':0.5,
                    'efficiency':100,
                    'polarization':'RHCP',
                    'circ2lambda_ratio':1,
                    'step2lmabda_ratio':0.23,
                    'no_turns':5,
                    'helix_type': 'Hansen-Woodyard'
                    }
    # Assign inputs to replace defaults
    parameters = includes.misc.update_params(parameters,input_parameters)

    return parameters

def __helix_pattern(vth, vph, k, wavelength, parameters):
    circumference = parameters['circ2lambda_ratio']*wavelength
    diameter = circumference/np.pi
    spacing = parameters['step2lmabda_ratio']*wavelength
    l_0 = np.sqrt(spacing**2 + circumference**2)

    if parameters['helix_type'] == 'Hansen-Woodyard':
        p_mult = (l_0/wavelength)/((spacing/wavelength) +\
                               (2*parameters['no_turns'] + 1)/(2*parameters['no_turns']))
        #print(p_mult)
    elif parameters['helix_type'] == 'ordinary':
        p_mult = (l_0/wavelength) / ((spacing/wavelength) + 1)

    h_psi = k*(spacing*np.cos(vth) - l_0/p_mult)

    e_th_vec = (np.sin(np.pi/(2*parameters['no_turns'])))*np.cos(vth)*\
        (np.sin((parameters['no_turns']/2)*h_psi) / np.sin(0.5*h_psi))

    e_th = np.tile(e_th_vec, vph.__len__()-1)
    e_th = np.reshape(e_th, (vth.__len__(),vph.__len__()-1) ,'F')
    e_ph = copy.deepcopy(e_th)

    if parameters['polarization'] == 'RHCP':
        e_th = e_th * np.exp(-1j*k*np.pi/2)
        e_ph = e_ph * np.exp(-1j*k)
    elif parameters['polarization'] == 'LHCP':
        e_th = e_th * np.exp(-1j*k)
        e_ph = e_ph * np.exp(-1j*k*np.pi/2)

    print('''{} Helix pattern generated:   circumference [m]: {};
                                diameter [m]: {};
                                spacing [m]: {};
                                number of turns [#]: {};
                                total length [m]: {}
                                expected AR [dB]: {}''' \
            .format(parameters['polarization'], \
                    circumference,\
                    diameter,\
                    spacing,\
                    parameters['no_turns'],\
                    spacing*parameters['no_turns'],\
                    10*np.log10((2*parameters['no_turns'] + 1)/(2*parameters['no_turns']))))

    return e_th, e_ph

def generate_analytical_pattern(patterns_obj, input_parameters = None):
    '''
    Creates an analytical antenna pattern.

    Args:
        patterns_obj (:class:`~includes.patterns.Patterns`): an instance of class \
            :class:`~includes.patterns.Patterns`
        input_parameters (dict): python dictionary of parameters. See \
            :func:`~includes.data_load_store.get_analytical_parameters` for details.

    Returns:
        The new pattern is appended to the list of patterns.
    '''
    # Define some constants first
    constants =     {   'speed_of_light': 299792458, # [m/s] speed of light in vacuum
                        'eta': 120*np.pi, # free space impedance
                        'i_0': 1, # constant
                        'radius': 1 # reference distance
                    }
    parameters = get_analytical_parameters(input_parameters)

    wavelength = constants['speed_of_light']/parameters['frequency']
    k = 2*np.pi/wavelength # Wavenumber
    vth = np.array(range(0,180+parameters['resolution'][0],parameters['resolution'][0]))*np.pi/180
    # vph is defined 0:360 for the integration only. For normal matrix storage the 360 degrees
    # is not needed
    vph = np.array(range(0, 360+parameters['resolution'][1], parameters['resolution'][1]))*np.pi/180

    if parameters['ant_type'] == 'dipole':
        e_th_vec = ((1j*constants['eta']*constants['i_0']*np.exp(-1j*(k*constants['radius'])))/ \
                        (2*np.pi*constants['radius']))*\
                        ((np.cos(np.cos(vth)*(k*parameters['length']*wavelength)/2)-\
                        np.cos((k*parameters['length']*wavelength)/2))/(np.sin(vth)))
        e_th_vec = np.nan_to_num(e_th_vec)

        e_th = np.tile(e_th_vec, vph.__len__()-1)
        e_th = np.reshape(e_th, (vth.__len__(),vph.__len__()-1) ,'F')
        e_ph = np.zeros((vth.__len__(),vph.__len__()-1), dtype = complex)
    elif parameters['ant_type'] == 'point source':
        e_th = np.ones((vth.__len__(),vph.__len__()-1))*\
                np.exp(1j*np.random.rand(vth.__len__(),vph.__len__()-1)*2*np.pi)
        e_ph = np.ones((vth.__len__(),vph.__len__()-1))*\
                np.exp(1j*np.random.rand(vth.__len__(),vph.__len__()-1)*2*np.pi)
    elif parameters['ant_type'] == 'delta function':
        e_th = np.zeros((vth.__len__(),vph.__len__()-1), dtype=complex)
        e_ph = e_th
        e_th[(vth.__len__()-1)//2,0] = 1*np.exp(1j*2*np.pi*np.random.rand(1))
        e_ph[(vth.__len__()-1)//2,0] = 1*np.exp(1j*2*np.pi*np.random.rand(1))
    elif parameters['ant_type'] == 'helix':
        e_th, e_ph = __helix_pattern(vth, vph, k, wavelength, parameters)
    else:
        sys.exit("Unknown analytical antenna type!")
    # Normalize the power
    g_abs = np.abs(e_th)**2 + np.abs(e_ph)**2 # compute current total power
    for i in range(0, int(180//parameters['resolution'][0]+1)):
        g_abs[i,:] = g_abs[i, :] * np.sin(vth[i])
    # Close the sphere:
    g_abs = np.append(g_abs, np.transpose([g_abs[:, 0]]), axis=1)
    g_norm = (parameters['efficiency']/100)*((4*np.pi)/(np.trapz(np.trapz(g_abs, vth, axis=0), \
                                                                 vph)))
    e_th = e_th*np.sqrt(g_norm)
    e_ph = e_ph*np.sqrt(g_norm)

    # Store the generated pattern
    new_pat = get_p_struct()
    new_pat['source'] = 'Analytical ' + parameters['ant_type']
    new_pat['file'] = 'N/A'
    new_pat['frequency'] = parameters['frequency']
    new_pat['port'] = 1
    new_pat['th_step'] = parameters['resolution'][0]
    new_pat['ph_step'] = parameters['resolution'][1]
    new_pat['ff'] = np.stack((e_th, e_ph))

    patterns_obj.patterns.append(new_pat)

    return patterns_obj

def save_to_excel(patterns_obj, pat_inds, field = 'Gabs', \
                        field_format = 'dB',filename = 'Default_XLS'):
    '''
    Saves selected patterns to an Excel file

    Args:
        patterns_obj (object): object of class Patterns. See\
                :class:`~includes.patterns`
        pat_inds (list): List of integers of the patterns to be stored
        field (str): string describing the field to be stored. See\
                :func:`~includes.field_operations.convert_field`
        field_format (str): field format of the field to be stored. See\
                :func:`~includes.field_operations.convert_field_format`
        filename (str): Filename to use. The file will be stored in the output folder. If a file of\
            the same name exists new sheets will be added. If the file does not exist it will be\
            created.
    '''
    if not os.path.exists('output'):
        os.makedirs('output')
    filename = 'output//'+filename+'.xlsx'
    if os.path.isfile(filename):
        workbook = load_workbook(filename)
    else:
        workbook = Workbook()

    for pat in pat_inds:
        temp_field,_,context = patterns_obj.fetch_field_with_context(pat, field, field_format)

        worksheet = workbook.create_sheet(title=str(pat)+'_'+field+'_'+field_format)

        __fill_worksheet(worksheet, context, temp_field)

    workbook.save(filename)

def __fill_worksheet(worksheet, context, data):
    '''
    helper function for excel file storage
    '''
    row_pad = 1
    col_pad = 1

    for key, value in context.items():
        worksheet.cell(row_pad, col_pad, key)
        worksheet.cell(row_pad, col_pad+1, str(value))
        row_pad = row_pad + 1

    vth,vph = includes.misc.get_angle_vectors(data, 'open', return_mesh = False)
    vth = vth*180/np.pi
    vph = vph*180/np.pi

    col_pad = col_pad + 2

    worksheet.cell(row_pad+1, col_pad+1, 'Phi angles [deg]')
    worksheet.merge_cells(start_row=row_pad+1, start_column=col_pad+1, end_row=row_pad+1, \
                   end_column=col_pad+1+data.shape[1]-1)
    row_pad = row_pad + 1

    for i,value in enumerate(vph):
        worksheet.cell(row_pad+1, col_pad+1+i, value)
    row_pad = row_pad + 1

    worksheet.cell(row_pad+2, 1, 'Theta angles [deg]')
    worksheet.merge_cells(start_row=row_pad+2, start_column=1, \
                   end_row=row_pad+2+data.shape[0]-1, end_column=1)

    for i,value in enumerate(vth):
        worksheet.cell(row_pad+2+i, 2, value)

    row_pad = row_pad + 2
    col_pad = col_pad + 1

    for pat_row in range(data.shape[0]):
        for pat_col in range(data.shape[1]):
            worksheet.cell(pat_row+row_pad, pat_col+col_pad, data[pat_row,pat_col])

def save_to_stk(patterns_obj, pat_inds, field = 'Gabs'):
    '''
    Saves selected patterns for usage in STK - separate text file for each pattern. Filenames \
    match the pattern's description string.

    Args:
        patterns_obj (object): object of class Patterns. See\
                :class:`~includes.patterns`
        pat_inds (list): List of integers of the patterns to be stored
        field (str): string describing the field to be stored. See\
                :func:`~includes.field_operations.convert_field`. Note that STK always uses dB.
    '''
    if not os.path.exists('output'):
        os.makedirs('output')

    for pat in pat_inds:
        context = patterns_obj.get_context(pat)
        legend = (str(context['Frequency [Hz]'] / 1E6) + " MHz @Port " + \
                        str(context['Port [#]']))

        filename = 'output//'+'STK Export - '+field+' - ' + legend +'.txt'

        data,_ = patterns_obj.fetch_field(pat,field, field_format='dB')

        vth,vph = includes.misc.get_angle_vectors(data, 'open', return_mesh = False)
        vth = np.round(vth*180/np.pi,0)
        vph = np.round(vph*180/np.pi,0)

        file = open(filename, 'wt')

        # Write some header
        file.write("stk.v.11.2.0\n")
        file.write("ThetaPhiPattern\n")
        file.write("AngleUnits Degrees\n")
        file.write("NumberOfPoints %d\n" % data.size)
        file.write("PatternData \n")
        for i_ph, ph in enumerate(vph):
            for i_th,th in enumerate(vth):
                file.write("%d\t%d\t%.3f\n" % (th,ph,data[i_th][i_ph]))

        file.close()
